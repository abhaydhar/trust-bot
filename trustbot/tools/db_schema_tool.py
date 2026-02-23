"""Tool for retrieving database schema from PostgreSQL or flat files."""

from __future__ import annotations

import csv
import io
import json
import logging
from pathlib import Path
from typing import Any

from trustbot.models.db_entity import DatabaseColumn, DatabaseTable

logger = logging.getLogger("trustbot.tools.db_schema")


# ---------------------------------------------------------------------------
# PostgreSQL live connection
# ---------------------------------------------------------------------------

async def fetch_pg_schema(
    host: str,
    port: int,
    database: str,
    schema: str,
    username: str,
    password: str,
    timeout: int = 30,
) -> list[DatabaseTable]:
    """
    Connect to PostgreSQL and retrieve all tables and columns for the
    given schema using information_schema.
    """
    import psycopg

    conninfo = psycopg.conninfo.make_conninfo(
        host=host,
        port=port,
        dbname=database,
        user=username,
        password=password,
        connect_timeout=timeout,
    )

    tables_query = """
        SELECT table_name
        FROM information_schema.tables
        WHERE table_schema = %s
          AND table_type = 'BASE TABLE'
        ORDER BY table_name
    """

    columns_query = """
        SELECT
            table_name,
            column_name,
            data_type,
            is_nullable,
            CASE
                WHEN kcu.column_name IS NOT NULL THEN true
                ELSE false
            END AS is_primary_key
        FROM information_schema.columns c
        LEFT JOIN information_schema.table_constraints tc
            ON tc.table_schema = c.table_schema
            AND tc.table_name = c.table_name
            AND tc.constraint_type = 'PRIMARY KEY'
        LEFT JOIN information_schema.key_column_usage kcu
            ON kcu.constraint_name = tc.constraint_name
            AND kcu.table_schema = tc.table_schema
            AND kcu.table_name = tc.table_name
            AND kcu.column_name = c.column_name
        WHERE c.table_schema = %s
        ORDER BY c.table_name, c.ordinal_position
    """

    async with await psycopg.AsyncConnection.connect(conninfo) as conn:
        async with conn.cursor() as cur:
            await cur.execute(tables_query, (schema,))
            table_rows = await cur.fetchall()

            await cur.execute(columns_query, (schema,))
            column_rows = await cur.fetchall()

    table_names = [row[0] for row in table_rows]
    columns_by_table: dict[str, list[DatabaseColumn]] = {t: [] for t in table_names}

    for row in column_rows:
        tbl, col_name, dtype, nullable_str, is_pk = row
        if tbl in columns_by_table:
            columns_by_table[tbl].append(DatabaseColumn(
                name=col_name,
                data_type=dtype or "",
                is_nullable=nullable_str == "YES",
                is_primary_key=bool(is_pk),
            ))

    result: list[DatabaseTable] = []
    for tbl in table_names:
        result.append(DatabaseTable(
            name=tbl,
            schema_name=schema,
            columns=columns_by_table.get(tbl, []),
        ))

    logger.info(
        "Fetched %d tables with %d total columns from %s:%s/%s schema=%s",
        len(result),
        sum(len(t.columns) for t in result),
        host, port, database, schema,
    )
    return result


# ---------------------------------------------------------------------------
# Flat file parsers
# ---------------------------------------------------------------------------

def parse_schema_file(filename: str, content: bytes) -> list[DatabaseTable]:
    """
    Auto-detect format by file extension and parse into DatabaseTable list.
    Supports CSV, JSON, and Excel (.xlsx).
    """
    ext = Path(filename).suffix.lower()
    if ext == ".csv":
        return _parse_csv(content)
    elif ext == ".json":
        return _parse_json(content)
    elif ext == ".xlsx":
        return _parse_excel(content)
    else:
        raise ValueError(
            f"Unsupported file format: {ext}. Use .csv, .json, or .xlsx."
        )


def _parse_csv(content: bytes) -> list[DatabaseTable]:
    """Parse CSV with columns: table_name, column_name, data_type, is_nullable, is_primary_key."""
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))

    tables: dict[str, list[DatabaseColumn]] = {}
    for row in reader:
        table_name = (row.get("table_name") or "").strip()
        if not table_name:
            continue
        col = DatabaseColumn(
            name=(row.get("column_name") or "").strip(),
            data_type=(row.get("data_type") or "").strip(),
            is_nullable=_parse_bool(row.get("is_nullable", "true")),
            is_primary_key=_parse_bool(row.get("is_primary_key", "false")),
        )
        tables.setdefault(table_name, []).append(col)

    return [
        DatabaseTable(name=name, columns=cols)
        for name, cols in tables.items()
    ]


def _parse_json(content: bytes) -> list[DatabaseTable]:
    """Parse JSON with structure: {"tables": [{"name": ..., "columns": [...]}]}."""
    data = json.loads(content.decode("utf-8"))

    raw_tables: list[dict[str, Any]]
    if isinstance(data, dict) and "tables" in data:
        raw_tables = data["tables"]
    elif isinstance(data, list):
        raw_tables = data
    else:
        raise ValueError(
            'JSON must be {"tables": [...]} or a list of table objects.'
        )

    result: list[DatabaseTable] = []
    for tbl in raw_tables:
        columns = []
        for col in tbl.get("columns", []):
            columns.append(DatabaseColumn(
                name=col.get("name", ""),
                data_type=col.get("data_type", ""),
                is_nullable=col.get("is_nullable", True),
                is_primary_key=col.get("is_primary_key", False),
            ))
        result.append(DatabaseTable(
            name=tbl.get("name", ""),
            schema_name=tbl.get("schema_name", ""),
            columns=columns,
        ))
    return result


def _parse_excel(content: bytes) -> list[DatabaseTable]:
    """Parse Excel (.xlsx) with same columns as CSV format."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no active sheet.")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(h or "").strip().lower() for h in rows[0]]
    col_map = {name: idx for idx, name in enumerate(header)}

    def _get(row: tuple, key: str) -> str:
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return ""
        return str(row[idx] or "").strip()

    tables: dict[str, list[DatabaseColumn]] = {}
    for row in rows[1:]:
        table_name = _get(row, "table_name")
        if not table_name:
            continue
        col = DatabaseColumn(
            name=_get(row, "column_name"),
            data_type=_get(row, "data_type"),
            is_nullable=_parse_bool(_get(row, "is_nullable") or "true"),
            is_primary_key=_parse_bool(_get(row, "is_primary_key") or "false"),
        )
        tables.setdefault(table_name, []).append(col)

    wb.close()
    return [
        DatabaseTable(name=name, columns=cols)
        for name, cols in tables.items()
    ]


def _parse_bool(value: str) -> bool:
    return str(value).strip().lower() in ("true", "yes", "1", "t", "y")
